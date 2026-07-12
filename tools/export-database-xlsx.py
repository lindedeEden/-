# -*- coding: utf-8 -*-
"""Export APP_DATABASE from database.js to Excel (headers as columns, rows as records)."""
import json
import os
import sys

import js2py
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_JS = os.path.join(ROOT, 'assets', 'data', 'database.js')
OUT_XLSX = os.path.join(ROOT, 'database-export.xlsx')


def load_database():
    with open(DB_JS, 'r', encoding='utf-8') as f:
        content = f.read()
    static_part = content.split('(function ensureOtherFindingsFields')[0].strip()
    if static_part.endswith(';'):
        static_part = static_part[:-1]
    ctx = js2py.EvalJs()
    ctx.execute(static_part)
    db = ctx.APP_DATABASE.to_dict()
    apply_runtime_defaults(db)
    return db


def apply_runtime_defaults(db):
    """Mirror post-load IIFEs in database.js for complete export."""
    specimens = db.get('specimens') or []
    other_keys = ['nrbc', 'giantPlt', 'megakaryocyte', 'smudgeCell', 'artefact']
    copy_keys = ['band', 'segmentedNeutrophil', 'eosinophil', 'monocyte', 'basophil', 'lymphocyte']
    empty_keys = [
        'atypicalLymphocyte', 'blast', 'promyelocyte', 'myelocyte', 'metamyelocyte',
        'hypersegmented', 'promonocyte', 'plasmaCell', 'abnormalLymphocyte',
    ]

    for s in specimens:
        s.setdefault('metrics', {})
        s.setdefault('prevReport', {})
        for k in other_keys:
            s['metrics'].setdefault(k, '-')
            s['prevReport'].setdefault(k, '-')

        if not s.get('prevReportDate') and s.get('analysisTime'):
            import re
            from datetime import datetime, timedelta
            m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', str(s['analysisTime']))
            if m:
                at = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                s['prevReportDate'] = (at - timedelta(days=7)).strftime('%Y-%m-%d')

        flow = s.setdefault('flowCyt', {})
        metrics = s['metrics']
        for k in copy_keys:
            flow.setdefault(k, metrics.get(k, '-'))
        for k in empty_keys:
            flow[k] = '-'

        cbc = s.setdefault('cbc', {})
        cbc.setdefault('wbc', metrics.get('wbc', '-'))
        cbc.setdefault('plt', metrics.get('plt', '-'))
        cbc.setdefault('rbc', '4.20')
        cbc.setdefault('hb', '13.5')
        cbc.setdefault('hct', '40.0')
        cbc.setdefault('mcv', '90.0')
        cbc.setdefault('mch', '30.0')
        cbc.setdefault('mchc', '33.0')


def flatten_value(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, list):
        if not value:
            return ''
        if all(isinstance(x, (str, int, float, bool)) or x is None for x in value):
            return '; '.join('' if x is None else str(x) for x in value)
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def flatten_record(record, prefix=''):
    row = {}
    for key, value in record.items():
        col = f'{prefix}{key}' if not prefix else f'{prefix}.{key}'
        if isinstance(value, dict):
            row.update(flatten_record(value, col))
        else:
            row[col] = flatten_value(value)
    return row


def collect_columns(rows):
    cols = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                cols.append(key)
    return cols


def write_sheet(ws, title, rows, columns=None):
    ws.title = title
    if not rows:
        ws.append(['(無資料)'])
        return
    columns = columns or collect_columns(rows)
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in rows:
        ws.append([row.get(col, '') for col in columns])
    for col_idx, col in enumerate(columns, 1):
        max_len = len(str(col))
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 40))


def main():
    db = load_database()
    wb = Workbook()

    specimens = db.get('specimens') or []
    specimen_rows = [flatten_record(s) for s in specimens]
    ws_specimens = wb.active
    write_sheet(ws_specimens, 'specimens', specimen_rows)

    departments = db.get('departments') or []
    if departments:
        ws_dept = wb.create_sheet('departments')
        ws_dept.append(['department'])
        ws_dept.cell(1, 1).fill = PatternFill('solid', fgColor='4472C4')
        ws_dept.cell(1, 1).font = Font(bold=True, color='FFFFFF')
        for d in departments:
            ws_dept.append([d])
        ws_dept.column_dimensions['A'].width = 24

    machines = db.get('machines') or []
    if machines:
        ws_machine = wb.create_sheet('machines')
        ws_machine.append(['machine'])
        ws_machine.cell(1, 1).fill = PatternFill('solid', fgColor='4472C4')
        ws_machine.cell(1, 1).font = Font(bold=True, color='FFFFFF')
        for m in machines:
            ws_machine.append([m])
        ws_machine.column_dimensions['A'].width = 12

    wb.save(OUT_XLSX)
    print(f'Exported {len(specimen_rows)} specimens to: {OUT_XLSX}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
